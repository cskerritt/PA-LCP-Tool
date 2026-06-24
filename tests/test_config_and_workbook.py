"""Tests for config loading, the CLI sample path, and workbook generation."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

import palcp
from palcp.config import build_plan, load_assumptions, load_items
from palcp.economics import project
from palcp.pricing import apply_pricing, load_pricing
from palcp.validate import validate_plan
from palcp.workbook import save_workbook

DATA = Path(palcp.__file__).parent / "data"

EXPECTED_SHEETS = [
    "Cover", "Summary", "Current-Cost Tables", "PV Projection",
    "Annual Schedule", "Assumptions & Rates", "Data Sources",
    "Methodology & Standards", "Medical Foundation", "Validation",
]


def _build_sample_plan():
    plan = build_plan(
        load_assumptions(DATA / "sample_assumptions.yaml"),
        load_items(DATA / "sample_plan_items.csv"),
    )
    apply_pricing(plan.items, load_pricing(DATA / "sample_pricing.csv"))
    return plan


def test_sample_files_load_and_have_items():
    plan = _build_sample_plan()
    assert len(plan.items) == 14
    # Code-resolved items should now carry a positive unit cost.
    wheelchair = next(i for i in plan.items if i.item == "Manual wheelchair")
    assert wheelchair.unit_cost == 2480


def test_sample_plan_validates_clean():
    plan = _build_sample_plan()
    report = validate_plan(plan)
    assert report.ok, report.summary()


def test_workbook_has_all_sheets(tmp_path):
    plan = _build_sample_plan()
    report = validate_plan(plan)
    result = project(plan)
    out = tmp_path / "wb.xlsx"
    save_workbook(result, report, str(out), generated_on="2026-06-23 00:00")
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_workbook_totals_match_engine(tmp_path):
    plan = _build_sample_plan()
    result = project(plan)
    out = tmp_path / "wb.xlsx"
    save_workbook(result, validate_plan(plan), str(out),
                  generated_on="2026-06-23 00:00")
    wb = load_workbook(out)
    ws = wb["Summary"]
    # Find the TOTAL row and compare present value (col D) to the engine.
    total_pv = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "TOTAL":
            total_pv = row[3]
    assert total_pv == pytest.approx(round(result.grand_total_present_value, 2))


def test_misaligned_csv_reports_column(tmp_path):
    p = tmp_path / "bad.csv"
    # 'unit_cost' column contains text -> should raise a descriptive error.
    p.write_text(
        "category,item,unit_cost,frequency_per_year\n"
        "Cat,Thing,not_a_number,2\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError) as exc:
        load_items(p)
    assert "unit_cost" in str(exc.value)
    assert "row 2" in str(exc.value)


def test_build_plan_requires_age():
    cfg = {"life_expectancy": {"additional_years": 10}}
    with pytest.raises(ValueError):
        build_plan(cfg, [])


def test_cli_sample_smoke(tmp_path):
    from palcp.cli import main
    out = tmp_path / "cli_sample.xlsx"
    rc = main(["sample", "--out", str(out)])
    assert rc == 0
    assert out.exists()


def test_load_items_empty_xlsx_returns_empty(tmp_path):
    from openpyxl import Workbook
    p = tmp_path / "empty.xlsx"
    Workbook().save(p)  # a workbook with one empty sheet
    assert load_items(p) == []


def test_cli_pricing_path_with_preset_suffix(tmp_path):
    """`path:preset` only splits when the suffix is a known preset key."""
    from palcp.cli import _load_pricing_arg
    p = tmp_path / "vendor.csv"
    p.write_text("CPT/HCPCS,Description,Charge\n99214,Office visit,198\n",
                 encoding="utf-8")
    table = _load_pricing_arg([f"{p}:va_reasonable_charges"])
    assert table.by_code("99214")[0].source == "VA Reasonable Charges"
