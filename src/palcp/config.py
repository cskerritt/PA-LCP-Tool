"""Load a :class:`~palcp.models.Plan` from input files.

Inputs are deliberately boring and inspectable:

* **assumptions** -- a YAML file (case metadata, life expectancy, discount rate,
  growth rates).  See ``data/templates/assumptions.template.yaml``.
* **care items** -- a CSV or ``.xlsx`` with one row per recommended good/service.
  See ``data/templates/plan_items.template.csv``.

The built-in default growth/discount rates are intentionally tagged as
``PLACEHOLDER`` sources so the validator flags them until you substitute current,
cited figures as of the report date.  Nothing here invents an authoritative
number on your behalf.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Optional

import yaml

from .models import CareItem, Claimant, DiscountRate, GrowthRate, LifeExpectancy, Plan

_PLACEHOLDER = "PLACEHOLDER - confirm against current published data as of the report date"

# Growth-rate keys the tool understands out of the box.  Care items reference
# these via their ``growth_key`` column.
DEFAULT_GROWTH_KEYS = {
    "medical_services": "Medical Care Services",
    "rx": "Prescription Drugs",
    "dme": "Durable Medical Equipment / Supplies",
    "facility": "Hospital & Facility Services",
    "attendant_care_wage": "Attendant / Home-Health Care (wage-based)",
    "general": "General (CPI-U, all items)",
}


def default_growth_rates() -> dict[str, GrowthRate]:
    """Placeholder growth rates (all flagged for replacement).

    The *rates* below are conservative round numbers ONLY so the tool runs out
    of the box; the ``source`` field marks each as a placeholder so the
    validator will warn until you enter the figure and citation you will defend.
    """
    seed = {
        "medical_services": 0.030,
        "rx": 0.030,
        "dme": 0.020,
        "facility": 0.035,
        "attendant_care_wage": 0.030,
        "general": 0.024,
    }
    return {
        key: GrowthRate(
            key=key,
            label=label,
            annual_rate=seed[key],
            source=_PLACEHOLDER,
            note="Default placeholder rate; replace with a cited published series.",
        )
        for key, label in DEFAULT_GROWTH_KEYS.items()
    }


# --------------------------------------------------------------------------- #
# Small parsing helpers
# --------------------------------------------------------------------------- #
def _opt_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().replace("$", "").replace(",", "").replace("%", "")
    if s == "":
        return None
    return float(s)


def _opt_bool(value: Any) -> bool:
    return str(value).strip().lower() in ("1", "true", "yes", "y", "t")


def _opt_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


# --------------------------------------------------------------------------- #
# Care items
# --------------------------------------------------------------------------- #
_ITEM_FIELDS = (
    "category", "item", "description", "code", "code_type", "pricing_source",
    "percentile", "geographic_basis", "retrieval_date", "unit_cost",
    "units_per_occurrence", "frequency_per_year", "every_n_years", "one_time",
    "one_time_age", "start_age", "end_age", "growth_key", "medical_foundation",
    "notes",
)


def _item_from_row(row: dict[str, Any], row_number: int = 0) -> CareItem:
    g = {k.lower().strip(): v for k, v in row.items() if k is not None}

    def val(name: str):
        return g.get(name)

    def num(name: str):
        """Parse a numeric field, naming the column/row on failure.

        A bad value here almost always means the row's columns are misaligned
        (e.g. a missing comma), so the error points straight at the cause.
        """
        try:
            return _opt_float(val(name))
        except ValueError:
            where = f" (row {row_number})" if row_number else ""
            raise ValueError(
                f"Could not read a number from column {name!r}{where}: "
                f"got {val(name)!r}. Check for a missing/extra comma or a "
                f"value in the wrong column."
            ) from None

    return CareItem(
        category=_opt_str(val("category")) or "Uncategorized",
        item=_opt_str(val("item")),
        description=_opt_str(val("description")),
        code=_opt_str(val("code")),
        code_type=_opt_str(val("code_type")),
        pricing_source=_opt_str(val("pricing_source")),
        percentile=num("percentile"),
        geographic_basis=_opt_str(val("geographic_basis")),
        retrieval_date=_opt_str(val("retrieval_date")),
        unit_cost=num("unit_cost") or 0.0,
        units_per_occurrence=num("units_per_occurrence") or 1.0,
        frequency_per_year=num("frequency_per_year"),
        every_n_years=num("every_n_years"),
        one_time=_opt_bool(val("one_time")),
        one_time_age=num("one_time_age"),
        start_age=num("start_age"),
        end_age=num("end_age"),
        growth_key=_opt_str(val("growth_key")) or "medical_services",
        medical_foundation=_opt_str(val("medical_foundation")),
        notes=_opt_str(val("notes")),
    )


def load_items(path: str | Path, sheet: Optional[str] = None) -> list[CareItem]:
    """Load care items from a CSV or ``.xlsx`` file."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        items = [
            _item_from_row(dict(zip(header, r)), row_number=n)
            for n, r in enumerate(rows, start=2)
            if any(c not in (None, "") for c in r)
        ]
        wb.close()
        return items

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return [
            _item_from_row(r, row_number=n)
            for n, r in enumerate(reader, start=2)
            if any(v not in (None, "") for v in r.values())
        ]


# --------------------------------------------------------------------------- #
# Assumptions -> Plan
# --------------------------------------------------------------------------- #
def load_assumptions(path: str | Path) -> dict:
    with open(path, encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    if not isinstance(data, dict):
        raise ValueError(f"Assumptions file {path} did not parse to a mapping.")
    return data


def build_plan(config: dict, items: list[CareItem]) -> Plan:
    """Assemble a :class:`Plan` from a parsed assumptions mapping + care items."""
    c = config.get("claimant", {}) or {}
    claimant = Claimant(
        name=_opt_str(c.get("name")),
        dob=_opt_str(c.get("dob")),
        sex=_opt_str(c.get("sex")) or "total",
        age_at_report=_opt_float(c.get("age_at_report")),
        residence=_opt_str(c.get("residence")),
        notes=_opt_str(c.get("notes")),
    )

    le = config.get("life_expectancy", {}) or {}
    age_at_report = (
        _opt_float(le.get("age_at_report"))
        if le.get("age_at_report") is not None
        else claimant.age_at_report
    )
    if age_at_report is None:
        raise ValueError(
            "age_at_report must be set (claimant.age_at_report or "
            "life_expectancy.age_at_report)."
        )
    life_expectancy = LifeExpectancy(
        age_at_report=age_at_report,
        additional_years=_opt_float(le.get("additional_years")) or 0.0,
        source=_opt_str(le.get("source")),
        citation_url=_opt_str(le.get("citation_url")),
        as_of=_opt_str(le.get("as_of")),
        note=_opt_str(le.get("note")),
    )

    dr = config.get("discount_rate", {}) or {}
    discount = DiscountRate(
        annual_rate=_opt_float(dr.get("annual_rate")) or 0.0,
        basis=_opt_str(dr.get("basis")) or "nominal",
        timing=_opt_str(dr.get("timing")) or "mid_year",
        source=_opt_str(dr.get("source")),
        citation_url=_opt_str(dr.get("citation_url")),
        as_of=_opt_str(dr.get("as_of")),
    )

    growth_rates = default_growth_rates()
    for key, spec in (config.get("growth_rates", {}) or {}).items():
        spec = spec or {}
        rate = _opt_float(spec.get("annual_rate"))
        label = _opt_str(spec.get("label")) or DEFAULT_GROWTH_KEYS.get(key, key)
        growth_rates[key] = GrowthRate(
            key=key,
            label=label,
            annual_rate=rate if rate is not None else
            (growth_rates[key].annual_rate if key in growth_rates else 0.0),
            source=_opt_str(spec.get("source")),
            citation_url=_opt_str(spec.get("citation_url")),
            as_of=_opt_str(spec.get("as_of")),
            note=_opt_str(spec.get("note")),
        )

    return Plan(
        claimant=claimant,
        life_expectancy=life_expectancy,
        discount_rate=discount,
        growth_rates=growth_rates,
        items=items,
        report_date=_opt_str(config.get("report_date")),
        base_year=int(config["base_year"]) if config.get("base_year") else None,
        evaluator=_opt_str(config.get("evaluator")),
        evaluator_credentials=_opt_str(config.get("evaluator_credentials")),
        jurisdiction=_opt_str(config.get("jurisdiction")),
        matter=_opt_str(config.get("matter")),
        percentile_policy=_opt_float(config.get("percentile_policy")) or 80.0,
        collateral_source_note=_opt_str(config.get("collateral_source_note")),
        rounding=int(config.get("rounding", 2)),
    )


def load_plan(
    assumptions_path: str | Path,
    items_path: str | Path,
    items_sheet: Optional[str] = None,
) -> Plan:
    """Convenience: load assumptions + items and build a :class:`Plan`."""
    config = load_assumptions(assumptions_path)
    items = load_items(items_path, sheet=items_sheet)
    return build_plan(config, items)
