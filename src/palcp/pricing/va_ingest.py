"""Ingest the official VA Reasonable Charges v5.26 workbooks into a VADataset.

Reads the 19 ``v5-26_Table-*.xlsx`` files (direct-charge + RVU + GAAF + reference)
and normalizes them into the engine's :class:`~palcp.pricing.va_charges.VADataset`.
Also provides SQLite persistence so the dataset (which contains AMA-copyrighted CPT
content and must stay out of git) can be built once and loaded fast.

Category names differ slightly between Table S (conversion factors), Table G
(``Conversion Factor/GAAF Category``), and Table L (``... CF GAAF`` columns); they
are matched on a normalized alphanumeric key and stored under the canonical Table S
name so the engine's lookups line up.
"""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from .va_charges import VAChargeBasis, VADataset, normalize_zip3
from .schema import normalize_code


def _num(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip().replace("$", "").replace(",", "")
    if s in ("", "Blank", "N/A", "NA", "BR"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _catkey(s) -> str:
    """Normalized alphanumeric key for matching category labels across tables."""
    t = str(s or "").lower().replace("cf gaaf", "").replace("gaaf", "")
    return "".join(ch for ch in t if ch.isalnum())


def _rows(path: Path, sheet: Optional[str] = None) -> list[list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    out = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def _find(d: Path, letter: str) -> Optional[Path]:
    """Locate the workbook for a table letter.

    Files look like 'v5-26_Table-G.xlsx' or 'v5-26-Table-A (1).xlsx'. The negative
    lookahead ensures e.g. letter 'A' does not match a hypothetical 'Table-AB'.
    """
    pat = re.compile(rf"Table-{letter}(?![A-Za-z])")
    for p in sorted(d.glob("*.xlsx")):
        if pat.search(p.name):
            return p
    return None


def ingest_va_tables(directory: str | Path) -> VADataset:
    """Build a :class:`VADataset` from a directory of VA v5.26 workbooks."""
    d = Path(directory)
    ds = VADataset(version="v5.26", effective_date="2026-01-01")

    # ---- Table S: conversion factors (canonical category names) -------------
    s_rows = _rows(_find(d, "S"), "Table S")
    catname_by_key: dict[str, str] = {}
    for r in s_rows[5:]:
        name, cf = (str(r[1]).strip() if r[1] else ""), _num(r[2])
        if name and cf is not None:
            ds.cf[name] = cf
            catname_by_key[_catkey(name)] = name

    def canon(label) -> Optional[str]:
        return catname_by_key.get(_catkey(label))

    # ---- Table M: modifier charge factors -----------------------------------
    m = _find(d, "M")
    if m:
        for r in _rows(m, "Table M")[5:]:
            mod, fac = (str(r[0]).strip() if r[0] else ""), _num(r[2])
            if mod and fac is not None:
                ds.modifier[mod] = fac

    # ---- GAAF tables --------------------------------------------------------
    # Table L (professional): per-category CF GAAF columns, keyed by ZIP.
    L = _rows(_find(d, "L"), "Table L")
    Lhdr = L[4]
    for r in L[5:]:
        z = normalize_zip3(r[0])
        if not z or str(r[0]).strip().lower().startswith("nation"):
            continue
        for ci in range(1, len(Lhdr)):
            cat = canon(Lhdr[ci])
            val = _num(r[ci]) if ci < len(r) else None
            if cat and val is not None:
                ds.gaaf[("L", z, cat)] = val
    # Table P (outpatient facility): single factor.
    for r in _rows(_find(d, "P"), "Table P")[5:]:
        z, v = normalize_zip3(r[0]), _num(r[1]) if len(r) > 1 else None
        if z and v is not None and not str(r[0]).strip().lower().startswith("nation"):
            ds.gaaf[("P", z, None)] = v
    # Table Q (DME): Non-Drug, Drug.
    for r in _rows(_find(d, "Q"), "Table Q")[5:]:
        z = normalize_zip3(r[0])
        if not z or str(r[0]).strip().lower().startswith("nation"):
            continue
        if _num(r[1]) is not None:
            ds.gaaf[("Q", z, "Non-Drug")] = _num(r[1])
        if len(r) > 2 and _num(r[2]) is not None:
            ds.gaaf[("Q", z, "Drug")] = _num(r[2])

    # ---- Charge tables ------------------------------------------------------
    def add(b: VAChargeBasis) -> None:
        ds.bases.setdefault(normalize_code(b.code), []).append(b)

    # Table F — Outpatient Facility (direct charge x Table P).
    for r in _rows(_find(d, "F"), "Table F")[7:]:
        code, ch = (str(r[0]).strip() if r[0] else ""), _num(r[4]) if len(r) > 4 else None
        if code and ch is not None:
            add(VAChargeBasis(code=code, table="F", charge_type="Outpatient Facility",
                              description=str(r[1] or ""), charge=ch, gaaf_table="P",
                              gaaf_category=None, methodology=str(r[5] or "") if len(r) > 5 else ""))
    # Table K — DME (direct charge x Table Q by Non-Drug/Drug).
    for r in _rows(_find(d, "K"), "Table K")[6:]:
        code, ch = (str(r[0]).strip() if r[0] else ""), _num(r[4]) if len(r) > 4 else None
        if not code or ch is None:
            continue
        gcat = "Drug" if (len(r) > 5 and "drug" in str(r[5]).lower()
                          and "non" not in str(r[5]).lower()) else "Non-Drug"
        mod = str(r[1]).strip() if len(r) > 1 and str(r[1]).strip() != "Blank" else ""
        add(VAChargeBasis(code=code, table="K", charge_type="DME",
                          description=str(r[2] or "") if len(r) > 2 else "", charge=ch,
                          gaaf_table="Q", gaaf_category=gcat, modifier=mod))
    # Table G — Physician/Professional (RVU x CF x Table L).
    for r in _rows(_find(d, "G"), "Table G")[7:]:
        code = str(r[0]).strip() if r[0] else ""
        if not code:
            continue
        cat = canon(r[3]) if len(r) > 3 else None
        if not cat:
            continue
        add(VAChargeBasis(
            code=code, table="G", charge_type="Physician/Professional",
            description=str(r[2] or "") if len(r) > 2 else "",
            work_rvu=_num(r[5]) if len(r) > 5 else None,
            facility_pe_rvu=_num(r[6]) if len(r) > 6 else None,
            nonfacility_pe_rvu=_num(r[7]) if len(r) > 7 else None,
            total_expense_rvu=_num(r[8]) if len(r) > 8 else None,
            cf_category=cat, gaaf_table="L", gaaf_category=cat,
            methodology=str(r[9] or "") if len(r) > 9 else "",
            modifier=str(r[1]).strip() if len(r) > 1 and str(r[1]).strip() != "Blank" else ""))
    # Table J — Pathology/Lab (Total RVU x Pathology CF x Table L Pathology).
    for r in _rows(_find(d, "J"), "Table J")[7:]:
        code = str(r[0]).strip() if r[0] else ""
        cat = canon(r[3]) if len(r) > 3 else None
        rvu = _num(r[4]) if len(r) > 4 else None
        if code and cat and rvu is not None:
            add(VAChargeBasis(code=code, table="J", charge_type="Pathology/Lab",
                              description=str(r[2] or "") if len(r) > 2 else "",
                              total_expense_rvu=rvu, cf_category=cat, gaaf_table="L",
                              gaaf_category=cat,
                              methodology=str(r[5] or "") if len(r) > 5 else ""))
    return ds


# --------------------------------------------------------------------------- #
# SQLite persistence
# --------------------------------------------------------------------------- #
_SCHEMA = """
CREATE TABLE meta (k TEXT PRIMARY KEY, v TEXT);
CREATE TABLE basis (code TEXT, table_letter TEXT, charge_type TEXT, description TEXT,
  charge REAL, work_rvu REAL, facility_pe_rvu REAL, nonfacility_pe_rvu REAL,
  total_expense_rvu REAL, cf_category TEXT, gaaf_table TEXT, gaaf_category TEXT,
  methodology TEXT, status_indicator TEXT, modifier TEXT);
CREATE INDEX ix_basis_code ON basis(code);
CREATE TABLE cf (category TEXT PRIMARY KEY, factor REAL);
CREATE TABLE gaaf (gaaf_table TEXT, zip3 TEXT, category TEXT, factor REAL);
CREATE INDEX ix_gaaf ON gaaf(gaaf_table, zip3, category);
CREATE TABLE modifier (modifier TEXT PRIMARY KEY, factor REAL);
"""


def to_sqlite(ds: VADataset, path: str | Path) -> None:
    p = Path(path)
    if p.exists():
        p.unlink()
    p.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(p)
    con.executescript(_SCHEMA)
    con.execute("INSERT INTO meta VALUES ('version', ?)", (ds.version,))
    con.execute("INSERT INTO meta VALUES ('effective_date', ?)", (ds.effective_date,))
    con.executemany("INSERT INTO basis VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", [
        (b.code, b.table, b.charge_type, b.description, b.charge, b.work_rvu,
         b.facility_pe_rvu, b.nonfacility_pe_rvu, b.total_expense_rvu, b.cf_category,
         b.gaaf_table, b.gaaf_category, b.methodology, b.status_indicator, b.modifier)
        for bs in ds.bases.values() for b in bs])
    con.executemany("INSERT INTO cf VALUES (?,?)", list(ds.cf.items()))
    con.executemany("INSERT INTO gaaf VALUES (?,?,?,?)",
                    [(t, z, c, v) for (t, z, c), v in ds.gaaf.items()])
    con.executemany("INSERT INTO modifier VALUES (?,?)", list(ds.modifier.items()))
    con.commit()
    con.close()


def from_sqlite(path: str | Path) -> VADataset:
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    meta = {r["k"]: r["v"] for r in con.execute("SELECT k, v FROM meta")}
    ds = VADataset(version=meta.get("version", ""),
                   effective_date=meta.get("effective_date", ""))
    for r in con.execute("SELECT * FROM basis"):
        ds.bases.setdefault(r["code"], []).append(VAChargeBasis(
            code=r["code"], table=r["table_letter"], charge_type=r["charge_type"],
            description=r["description"], charge=r["charge"], work_rvu=r["work_rvu"],
            facility_pe_rvu=r["facility_pe_rvu"], nonfacility_pe_rvu=r["nonfacility_pe_rvu"],
            total_expense_rvu=r["total_expense_rvu"], cf_category=r["cf_category"],
            gaaf_table=r["gaaf_table"], gaaf_category=r["gaaf_category"],
            methodology=r["methodology"], status_indicator=r["status_indicator"],
            modifier=r["modifier"]))
    for r in con.execute("SELECT * FROM cf"):
        ds.cf[r["category"]] = r["factor"]
    for r in con.execute("SELECT * FROM gaaf"):
        ds.gaaf[(r["gaaf_table"], r["zip3"], r["category"])] = r["factor"]
    for r in con.execute("SELECT * FROM modifier"):
        ds.modifier[r["modifier"]] = r["factor"]
    con.close()
    return ds
