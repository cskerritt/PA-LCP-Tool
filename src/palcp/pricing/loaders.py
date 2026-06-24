"""Loaders that normalise vendor pricing exports into a :class:`PricingTable`.

Two entry points cover almost everything:

* :func:`load_pricing` -- read a CSV or ``.xlsx`` using either the canonical
  column names or an explicit ``column_map`` you supply for an arbitrary export.
* :data:`PRESETS` -- ready-made ``column_map`` + defaults for common sources so
  you can do ``load_pricing("va.xlsx", preset="va_reasonable_charges")`` and only
  override what differs in your file.

Canonical columns (case-insensitive): ``source, code, code_type, description,
amount, percentile, geographic_area, effective_date, citation_url``.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Optional

from .schema import PriceRecord, PricingTable

CANONICAL_FIELDS = (
    "source",
    "code",
    "code_type",
    "description",
    "amount",
    "percentile",
    "geographic_area",
    "effective_date",
    "citation_url",
)

# Column-name presets for common vendor exports.  Map canonical field -> the
# header used in that vendor's file.  ``defaults`` fill fields the file omits.
PRESETS: dict[str, dict] = {
    "va_reasonable_charges": {
        "column_map": {
            "code": "CPT/HCPCS",
            "description": "Description",
            "amount": "Charge",
            "geographic_area": "Geographic Area",
        },
        "defaults": {
            "source": "VA Reasonable Charges",
            "code_type": "CPT/HCPCS",
            "citation_url": "https://www.va.gov/COMMUNITYCARE/revenue_ops/Reasonable_Charges.asp",
        },
    },
    "cms_dmepos": {
        "column_map": {
            "code": "HCPCS",
            "description": "DESCRIPTION",
            "amount": "FEE",
            "geographic_area": "STATE",
        },
        "defaults": {
            "source": "CMS DMEPOS Fee Schedule",
            "code_type": "HCPCS",
            "citation_url": "https://www.cms.gov/medicare/payment/fee-schedules/dmepos-fee-schedule",
        },
    },
    "mfus": {
        "column_map": {
            "code": "CPT",
            "description": "Description",
            "amount": "Amount",
            "percentile": "Percentile",
            "geographic_area": "ZIP",
        },
        "defaults": {
            "source": "Medical Fees in the United States (Context4/PMIC)",
            "code_type": "CPT",
        },
    },
}


def _to_float(value) -> Optional[float]:
    """Parse a currency/percent/number cell to float, or ``None``.

    Real fee schedules contain non-numeric placeholders ('N/A', 'BR'/'by
    report', 'see note') and percentiles written like '80%', plus space
    thousands separators ('1 200'). Returns ``None`` for blank/unparseable
    cells so the caller can skip the row rather than aborting the whole file.
    """
    if value is None:
        return None
    s = (
        str(value)
        .strip()
        .replace("$", "")
        .replace(",", "")
        .replace("%", "")
        .replace(" ", "")
    )
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_rows(path: str | Path, sheet: Optional[str]) -> tuple[list[str], Iterable[dict]]:
    """Yield (fieldnames, rows-as-dicts) from a CSV or xlsx file."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return [], []
        data = [dict(zip(header, r)) for r in rows_iter]
        wb.close()
        return header, data

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        header = list(reader.fieldnames or [])
        return header, list(reader)


def load_pricing(
    path: str | Path,
    *,
    preset: Optional[str] = None,
    column_map: Optional[dict[str, str]] = None,
    defaults: Optional[dict[str, str]] = None,
    sheet: Optional[str] = None,
) -> PricingTable:
    """Load a pricing file into a :class:`PricingTable`.

    Resolution order for each canonical field: ``column_map`` (explicit) ->
    preset's ``column_map`` -> a case-insensitive match against the canonical
    field name.  Missing values fall back to ``defaults`` then the preset's
    ``defaults``.
    """
    cmap: dict[str, str] = {}
    dflt: dict[str, str] = {}
    if preset:
        if preset not in PRESETS:
            raise ValueError(
                f"Unknown preset {preset!r}; available: {sorted(PRESETS)}"
            )
        cmap.update(PRESETS[preset].get("column_map", {}))
        dflt.update(PRESETS[preset].get("defaults", {}))
    if column_map:
        cmap.update(column_map)
    if defaults:
        dflt.update(defaults)

    header, rows = _read_rows(path, sheet)
    header_lower = {h.lower(): h for h in header}

    def resolve_header(field_name: str) -> Optional[str]:
        if field_name in cmap and cmap[field_name] in header:
            return cmap[field_name]
        if field_name in header_lower:
            return header_lower[field_name]
        return None

    resolved = {f: resolve_header(f) for f in CANONICAL_FIELDS}
    if resolved["code"] is None or (resolved["amount"] is None and "amount" not in dflt):
        raise ValueError(
            f"Could not locate required 'code' and 'amount' columns in {path}. "
            f"Found headers: {header}. Supply a column_map or use a preset."
        )

    records: list[PriceRecord] = []
    for raw in rows:
        def get(field_name: str):
            col = resolved[field_name]
            if col is not None and raw.get(col) not in (None, ""):
                return raw.get(col)
            return dflt.get(field_name)

        code = get("code")
        amount = _to_float(get("amount"))
        if code in (None, "") or amount is None:
            continue  # skip blank / non-priced rows
        records.append(
            PriceRecord(
                source=str(get("source") or "Unspecified source"),
                code=str(code),
                amount=amount,
                code_type=str(get("code_type") or ""),
                description=str(get("description") or ""),
                percentile=_to_float(get("percentile")),
                geographic_area=str(get("geographic_area") or ""),
                effective_date=str(get("effective_date") or ""),
                citation_url=str(get("citation_url") or ""),
            )
        )
    return PricingTable(records=records)


def load_many(specs: list[dict]) -> PricingTable:
    """Load and merge several pricing files.

    Each spec is a kwargs dict for :func:`load_pricing`, e.g.::

        load_many([
            {"path": "va.xlsx", "preset": "va_reasonable_charges"},
            {"path": "dmepos.csv", "preset": "cms_dmepos"},
        ])
    """
    table = PricingTable()
    for spec in specs:
        table.extend(load_pricing(**spec).records)
    return table
