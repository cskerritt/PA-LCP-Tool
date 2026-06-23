"""Shared openpyxl styling for a clean, conservative, court-ready workbook.

The palette is intentionally muted (navy headers, light-grey banding) -- the
goal is a document that reads as a professional exhibit, not a dashboard.
"""

from __future__ import annotations

import math

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

# --- Palette ------------------------------------------------------------- #
NAVY = "1F3864"
MID_BLUE = "2E5496"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
BAND_GREY = "FAFAFA"
RULE_GREY = "BFBFBF"
WHITE = "FFFFFF"
BLACK = "000000"

FONT_NAME = "Calibri"

# --- Number formats ------------------------------------------------------ #
FMT_MONEY = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_MONEY_CENTS = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"_);_(@_)'
FMT_PCT1 = "0.0%"
FMT_PCT2 = "0.00%"
FMT_INT = "#,##0"
FMT_NUM2 = "#,##0.00"

# --- Reusable parts ------------------------------------------------------ #
_thin = Side(style="thin", color=RULE_GREY)
_medium = Side(style="medium", color=NAVY)

BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_TOP = Border(top=Side(style="thin", color=BLACK))
BORDER_TOP_DOUBLE = Border(top=Side(style="double", color=BLACK))

TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, italic=True, color="404040")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
CATEGORY_FONT = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
BODY_FONT = Font(name=FONT_NAME, size=10, color=BLACK)
BODY_BOLD = Font(name=FONT_NAME, size=10, bold=True, color=BLACK)
SMALL_FONT = Font(name=FONT_NAME, size=8, italic=True, color="595959")
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True, color="404040")

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
CATEGORY_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
TOTAL_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
BAND_FILL = PatternFill("solid", fgColor=BAND_GREY)

LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def status_font(color: str) -> Font:
    """A bold font in an arbitrary hex color (used for severity/status cells)."""
    return Font(name=FONT_NAME, size=10, bold=True, color=color)


def set_widths(ws: Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws: Worksheet, title: str, subtitle: str, ncols: int) -> int:
    """Write a merged title + subtitle. Return the next free row (1-based)."""
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws.merge_cells(f"A2:{last_col}2")
        s = ws["A2"]
        s.value = subtitle
        s.font = SUBTITLE_FONT
        s.alignment = Alignment(horizontal="left", vertical="center")
        return 4
    return 3


def header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 30


def money(cell, cents: bool = False) -> None:
    cell.number_format = FMT_MONEY_CENTS if cents else FMT_MONEY
    cell.alignment = RIGHT
    cell.font = BODY_FONT


def note(ws: Worksheet, row: int, text: str, ncols: int) -> int:
    """Write a small footnote spanning ``ncols`` columns. Return next row.

    A merged cell does not auto-fit, so we estimate the wrapped height from the
    combined width of the spanned columns and set it explicitly; otherwise long
    notes are clipped to one line in Excel and in print.
    """
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = SMALL_FONT
    c.alignment = LEFT_WRAP

    total_width = 0.0
    for i in range(1, ncols + 1):
        w = ws.column_dimensions[get_column_letter(i)].width
        total_width += w if w else 8.43  # Excel default column width
    chars_per_line = max(20.0, total_width * 1.05)  # ~1 char per width unit
    lines = max(1, math.ceil(len(str(text)) / chars_per_line))
    ws.row_dimensions[row].height = max(15.0, 13.0 * lines + 3.0)
    return row + 1


def page_setup(ws: Worksheet, landscape: bool = True, repeat_header: int = 0) -> None:
    """Apply print setup: fit-to-width, optional landscape, repeating header."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = False
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    if repeat_header:
        ws.print_title_rows = f"1:{repeat_header}"
